from openpyxl import workbook

wb = workbook()

ws = wb.active
ws2 = wb.create_sheet("new_sheet2")
ws3 = wb.create_sheet("내 시트", 1)
wb.sheetnames
ws = wb["sheet"]
ws.title = "주소록"
print(ws.title)

ws['A1'] = "이름"
ws['B1'] = "전화번호"